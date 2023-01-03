import openpyxl

#  same data
# xlfile---> workbook--->sheets--->rows----> cells

# file="location of excel file"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["specify sheet name"]
#
# rows=sheet.max_row  # count number of rows in a excel sheet
# cols=sheet.max_column  # count number if columns in a excel sheet
#
# #  Reading all the rows & columns from excel sheet
#
# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(r,c).value,end="    ")
#     print()

# Multiple data

file="location of excel file"

workbook=openpyxl.load_workbook(file)
sheet=workbook.active     # (or) sheet=workbook["data] -- get active sheet from excel

sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value="engineer"

sheet.cell(2,1).value=456
sheet.cell(2,2).value="john"
sheet.cell(3,3).value="manager"


sheet.cell(3,1).value=896
sheet.cell(3,2).value="david"
sheet.cell(3,3).value="developer"

workbook.save(file)   # save the file after entering the data
